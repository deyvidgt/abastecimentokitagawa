# =================================================================
# EXPORT.PY — Exportação de relatórios (PDF e Excel)
# =================================================================
import logging
from datetime import datetime
from tkinter import messagebox

import pandas as pd

from config import C_ACCENT, C_SURFACE, C_BG, C_BORDER, C_TEXT, C_MUTED

# ── Dependências opcionais ────────────────────────────────────────
try:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors as rl_colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer)
    from reportlab.lib.enums import TA_CENTER, TA_LEFT   # noqa: F401
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False
    logging.warning("reportlab não instalado — export PDF desabilitado.")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, Reference
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    logging.warning("openpyxl não instalado — export Excel desabilitado.")


class ExportManager:

    # =============================================================
    # PDF
    # =============================================================
    @staticmethod
    def export_pdf(df: pd.DataFrame, filepath: str,
                   titulo: str = "Relatório de Frota"):
        if not HAS_REPORTLAB:
            messagebox.showerror("Erro",
                "Instale reportlab:\npip install reportlab")
            return False

        doc = SimpleDocTemplate(
            filepath, pagesize=landscape(A4),
            leftMargin=1.5*cm, rightMargin=1.5*cm,
            topMargin=1.5*cm, bottomMargin=1.5*cm)
        stls  = getSampleStyleSheet()
        elems = []

        title_style = ParagraphStyle(
            "T", parent=stls["Title"], fontSize=16,
            textColor=rl_colors.HexColor("#3B82F6"), spaceAfter=6)
        sub_style = ParagraphStyle(
            "S", parent=stls["Normal"], fontSize=9,
            textColor=rl_colors.HexColor("#64748B"), spaceAfter=14)

        elems.append(Paragraph(titulo, title_style))
        elems.append(Paragraph(
            f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')} · "
            f"{len(df)} registros", sub_style))

        # ── Sumário executivo ─────────────────────────────────────
        total_val  = df["valor"].sum()     if "valor"     in df.columns else 0
        total_ltrs = df["quantidade"].sum() if "quantidade" in df.columns else 0
        n_plac     = df["placa"].nunique()  if "placa"     in df.columns else 0
        media      = total_val / total_ltrs if total_ltrs > 0 else 0

        summ_data = [
            ["INVESTIMENTO TOTAL", "VOLUME (L)", "VEÍCULOS", "R$/LITRO"],
            [f"R$ {total_val:,.2f}", f"{total_ltrs:,.0f}",
             str(n_plac), f"R$ {media:.3f}"]
        ]
        summ_tbl = Table(summ_data, colWidths=[6*cm]*4)
        summ_tbl.setStyle(TableStyle([
            ("BACKGROUND",    (0,0),(-1,0), rl_colors.HexColor("#1A1D23")),
            ("TEXTCOLOR",     (0,0),(-1,0), rl_colors.HexColor("#64748B")),
            ("FONTNAME",      (0,0),(-1,0), "Helvetica-Bold"),
            ("FONTSIZE",      (0,0),(-1,0), 8),
            ("BACKGROUND",    (0,1),(-1,1), rl_colors.HexColor("#111318")),
            ("TEXTCOLOR",     (0,1),(-1,1), rl_colors.HexColor("#E2E8F0")),
            ("FONTNAME",      (0,1),(-1,1), "Helvetica-Bold"),
            ("FONTSIZE",      (0,1),(-1,1), 13),
            ("ALIGN",         (0,0),(-1,-1), "CENTER"),
            ("VALIGN",        (0,0),(-1,-1), "MIDDLE"),
            ("BOX",           (0,0),(-1,-1), 0.5, rl_colors.HexColor("#2A2D35")),
            ("INNERGRID",     (0,0),(-1,-1), 0.25, rl_colors.HexColor("#2A2D35")),
            ("TOPPADDING",    (0,0),(-1,-1), 8),
            ("BOTTOMPADDING", (0,0),(-1,-1), 8),
        ]))
        elems.append(summ_tbl)
        elems.append(Spacer(1, 0.5*cm))

        # ── Custo por veículo ─────────────────────────────────────
        if "placa" in df.columns and "valor" in df.columns:
            by_placa = df.groupby("placa").agg(
                total_valor=("valor", "sum"),
                total_litros=("quantidade", "sum"),
                n_abastecimentos=("valor", "count"),
            ).reset_index().sort_values("total_valor", ascending=False)

            h2 = ParagraphStyle(
                "H2", parent=stls["Heading2"], fontSize=11,
                textColor=rl_colors.HexColor("#E2E8F0"),
                spaceBefore=10, spaceAfter=4)
            elems.append(Paragraph("Custo por Veículo", h2))

            vh = [["Placa","Total R$","Litros","Registros"]]
            for _, r in by_placa.iterrows():
                vh.append([r["placa"], f"R$ {r['total_valor']:,.2f}",
                           f"{r['total_litros']:,.0f}",
                           str(int(r["n_abastecimentos"]))])
            vt = Table(vh, colWidths=[5*cm, 5*cm, 5*cm, 4*cm])
            vt.setStyle(TableStyle([
                ("BACKGROUND",    (0,0),(-1,0), rl_colors.HexColor("#3B82F6")),
                ("TEXTCOLOR",     (0,0),(-1,0), rl_colors.white),
                ("FONTNAME",      (0,0),(-1,0), "Helvetica-Bold"),
                ("FONTSIZE",      (0,0),(-1,0), 9),
                ("ROWBACKGROUNDS",(0,1),(-1,-1),
                    [rl_colors.HexColor("#1A1D23"), rl_colors.HexColor("#111318")]),
                ("TEXTCOLOR",     (0,1),(-1,-1), rl_colors.HexColor("#E2E8F0")),
                ("FONTSIZE",      (0,1),(-1,-1), 8),
                ("ALIGN",         (1,0),(-1,-1), "RIGHT"),
                ("ALIGN",         (0,0),(0,-1),  "LEFT"),
                ("BOX",           (0,0),(-1,-1), 0.5, rl_colors.HexColor("#2A2D35")),
                ("INNERGRID",     (0,0),(-1,-1), 0.25, rl_colors.HexColor("#2A2D35")),
                ("TOPPADDING",    (0,0),(-1,-1), 5),
                ("BOTTOMPADDING", (0,0),(-1,-1), 5),
            ]))
            elems.append(vt)
            elems.append(Spacer(1, 0.4*cm))

        # ── Detalhe (últimas 200 linhas) ──────────────────────────
        cols_show = [c for c in ["data","produto","placa","responsavel",
                                  "valor","quantidade","horimetro","categoria"]
                     if c in df.columns]
        df_show = df[cols_show].tail(200).copy()
        df_show.columns = [c.upper() for c in df_show.columns]

        elems.append(Paragraph("Registros Detalhados (últimos 200)", h2))
        col_w    = [25.5*cm / len(cols_show)] * len(cols_show)
        tbl_data = [df_show.columns.tolist()] + df_show.values.tolist()
        tbl = Table(tbl_data, colWidths=col_w, repeatRows=1)
        tbl.setStyle(TableStyle([
            ("BACKGROUND",    (0,0),(-1,0), rl_colors.HexColor("#3B82F6")),
            ("TEXTCOLOR",     (0,0),(-1,0), rl_colors.white),
            ("FONTNAME",      (0,0),(-1,0), "Helvetica-Bold"),
            ("FONTSIZE",      (0,0),(-1,0), 7),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),
                [rl_colors.HexColor("#1A1D23"), rl_colors.HexColor("#111318")]),
            ("TEXTCOLOR",     (0,1),(-1,-1), rl_colors.HexColor("#E2E8F0")),
            ("FONTSIZE",      (0,1),(-1,-1), 7),
            ("ALIGN",         (0,0),(-1,-1), "CENTER"),
            ("BOX",           (0,0),(-1,-1), 0.3, rl_colors.HexColor("#2A2D35")),
            ("INNERGRID",     (0,0),(-1,-1), 0.2,  rl_colors.HexColor("#2A2D35")),
            ("TOPPADDING",    (0,0),(-1,-1), 3),
            ("BOTTOMPADDING", (0,0),(-1,-1), 3),
        ]))
        elems.append(tbl)
        doc.build(elems)
        return True

    # =============================================================
    # Excel
    # =============================================================
    @staticmethod
    def export_excel(df: pd.DataFrame, filepath: str):
        if not HAS_OPENPYXL:
            messagebox.showerror("Erro",
                "Instale openpyxl:\npip install openpyxl")
            return False

        wb = openpyxl.Workbook()

        header_fill = PatternFill("solid", fgColor="3B82F6")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        data_font   = Font(color="E2E8F0", size=10)
        dark_fill   = PatternFill("solid", fgColor="1A1D23")
        alt_fill    = PatternFill("solid", fgColor="111318")
        thin = Border(
            left=Side(style="thin", color="2A2D35"),
            right=Side(style="thin", color="2A2D35"),
            top=Side(style="thin", color="2A2D35"),
            bottom=Side(style="thin", color="2A2D35"),
        )

        # ── Aba 1: Resumo ─────────────────────────────────────────
        ws1 = wb.active
        ws1.title = "Resumo"
        ws1.sheet_view.showGridLines = False
        ws1.column_dimensions["A"].width = 28
        ws1.column_dimensions["B"].width = 20

        ws1["A1"] = "RELATÓRIO CONSOLIDADO DE FROTA"
        ws1["A1"].font = Font(color="3B82F6", bold=True, size=14)
        ws1["A2"] = f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws1["A2"].font = Font(color="64748B", size=9)

        total_val  = df["valor"].sum()      if "valor"     in df.columns else 0
        total_ltrs = df["quantidade"].sum()  if "quantidade" in df.columns else 0
        n_plac     = df["placa"].nunique()   if "placa"     in df.columns else 0
        media      = total_val / total_ltrs  if total_ltrs > 0 else 0

        kpis = [
            ("INVESTIMENTO TOTAL", f"R$ {total_val:,.2f}"),
            ("VOLUME TOTAL (L)",   f"{total_ltrs:,.0f}"),
            ("FROTA ATIVA",        str(n_plac)),
            ("PREÇO MÉDIO / L",    f"R$ {media:.3f}"),
        ]
        for i, (k, v) in enumerate(kpis, start=4):
            ws1.cell(i, 1, k).fill  = header_fill
            ws1.cell(i, 1).font     = header_font
            ws1.cell(i, 1).border   = thin
            ws1.cell(i, 2, v).fill  = dark_fill
            ws1.cell(i, 2).font     = Font(color="E2E8F0", bold=True, size=12)
            ws1.cell(i, 2).border   = thin
            ws1.cell(i, 2).alignment = Alignment(horizontal="right")

        # Custo por veículo no resumo
        row = 10
        ws1.cell(row, 1, "CUSTO POR VEÍCULO").font = Font(
            color="E2E8F0", bold=True, size=11)
        row += 1
        for ci, col_title in enumerate(["Placa","Total R$","Litros","Registros"], 1):
            ws1.cell(row, ci, col_title).fill      = header_fill
            ws1.cell(row, ci).font                 = header_font
            ws1.cell(row, ci).border               = thin
            ws1.cell(row, ci).alignment            = Alignment(horizontal="center")

        if "placa" in df.columns:
            by_placa = df.groupby("placa").agg(
                total_valor=("valor", "sum"),
                total_litros=("quantidade", "sum"),
                n=("valor", "count"),
            ).reset_index().sort_values("total_valor", ascending=False)
            for ri, (_, r) in enumerate(by_placa.iterrows()):
                fill = dark_fill if ri % 2 == 0 else alt_fill
                row += 1
                for ci, v in enumerate(
                        [r["placa"], f"R$ {r['total_valor']:,.2f}",
                         f"{r['total_litros']:,.0f}", str(int(r["n"]))], 1):
                    ws1.cell(row, ci, v).fill    = fill
                    ws1.cell(row, ci).font        = data_font
                    ws1.cell(row, ci).border      = thin
                    ws1.cell(row, ci).alignment   = Alignment(
                        horizontal="right" if ci > 1 else "left")

        # ── Aba 2: Registros completos ────────────────────────────
        ws2 = wb.create_sheet("Registros")
        ws2.sheet_view.showGridLines = False
        cols_show = [c for c in ["data","produto","placa","responsavel","frota",
                                  "valor","quantidade","horimetro","categoria"]
                     if c in df.columns]
        for ci, col in enumerate(cols_show, 1):
            ws2.cell(1, ci, col.upper()).fill      = header_fill
            ws2.cell(1, ci).font                   = header_font
            ws2.cell(1, ci).border                 = thin
            ws2.cell(1, ci).alignment              = Alignment(horizontal="center")
            ws2.column_dimensions[
                openpyxl.utils.get_column_letter(ci)].width = 16
        for ri, (_, row_data) in enumerate(df[cols_show].iterrows(), 2):
            fill = dark_fill if ri % 2 == 0 else alt_fill
            for ci, val in enumerate(row_data, 1):
                ws2.cell(ri, ci, val).fill   = fill
                ws2.cell(ri, ci).font        = data_font
                ws2.cell(ri, ci).border      = thin
                ws2.cell(ri, ci).alignment   = Alignment(horizontal="center")

        # ── Aba 3: Mensal + gráfico ───────────────────────────────
        ws3 = wb.create_sheet("Mensal")
        ws3.sheet_view.showGridLines = False
        if "data" in df.columns:
            df2 = df.copy()
            df2["data"] = pd.to_datetime(df2["data"], errors="coerce")

            # BUG FIX: datas inválidas (NaT) causam "NaTType does not support
            # to_period" no pandas < 2.0. Mesmo no pandas ≥ 2.0, NaT periods
            # geram uma linha "NaT" inválida no Excel. O dropna remove essas
            # linhas antes do groupby, espelhando a correção já aplicada no
            # dashboard.py (Bug 3) e garantindo consistência entre ambos.
            df2 = df2.dropna(subset=["data"])

            df2["mes"] = df2["data"].dt.to_period("M")
            monthly = df2.groupby("mes")["valor"].sum().reset_index()
            monthly["mes"] = monthly["mes"].astype(str)

            for ci, h in enumerate(["Mês","Total R$"], 1):
                ws3.cell(1, ci, h).fill  = header_fill
                ws3.cell(1, ci).font     = header_font
                ws3.cell(1, ci).border   = thin
                ws3.cell(1, ci).alignment = Alignment(horizontal="center")
                ws3.column_dimensions[
                    openpyxl.utils.get_column_letter(ci)].width = 18
            for ri, (_, r) in enumerate(monthly.iterrows(), 2):
                fill = dark_fill if ri % 2 == 0 else alt_fill
                ws3.cell(ri,1,str(r["mes"])).fill = fill
                ws3.cell(ri,1).font  = data_font
                ws3.cell(ri,1).border = thin
                ws3.cell(ri,1).alignment = Alignment(horizontal="center")
                ws3.cell(ri,2,round(r["valor"],2)).fill = fill
                ws3.cell(ri,2).font  = data_font
                ws3.cell(ri,2).border = thin
                ws3.cell(ri,2).alignment = Alignment(horizontal="right")

            chart = BarChart()
            chart.type    = "col"
            chart.title   = "Custo Mensal"
            chart.y_axis.title = "R$"
            chart.x_axis.title = "Mês"
            chart.style   = 10
            chart.width   = 20
            chart.height  = 12
            data_ref = Reference(ws3, min_col=2, min_row=1,
                                 max_row=len(monthly)+1)
            cats     = Reference(ws3, min_col=1, min_row=2,
                                 max_row=len(monthly)+1)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats)
            ws3.add_chart(chart, "D2")

        wb.save(filepath)
        return True
