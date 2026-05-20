# =================================================================
# EXPORT_UTILS.PY — Exportação PDF e Excel com logo Kitagawa
# =================================================================
import io
import os
import pandas as pd
from datetime import datetime

# Caminhos possíveis do logo
_LOGO_PATHS = [
    "streamlit/logo.png",
    "logo.png",
    "../logo.png",
]

def _get_logo_path():
    for p in _LOGO_PATHS:
        if os.path.exists(p):
            return p
    return None


# ── Excel com logo ────────────────────────────────────────────────
def df_to_excel(df: pd.DataFrame, sheet_name: str = "Dados") -> bytes:
    return df_to_excel_multi({sheet_name: df})


def df_to_excel_multi(sheets: dict) -> bytes:
    """Gera Excel com múltiplas abas e logo no cabeçalho."""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.utils import get_column_letter

        header_fill = PatternFill("solid", fgColor="10B981")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        alt_fill    = PatternFill("solid", fgColor="1A1D23")
        dark_fill   = PatternFill("solid", fgColor="111318")
        data_font   = Font(color="E2E8F0", size=10)
        title_font  = Font(color="10B981", bold=True, size=14)
        sub_font    = Font(color="64748B", size=9)
        thin = Border(
            left=Side(style="thin",   color="2A2D35"),
            right=Side(style="thin",  color="2A2D35"),
            top=Side(style="thin",    color="2A2D35"),
            bottom=Side(style="thin", color="2A2D35"))

        logo_path = _get_logo_path()

        for sheet_name, df in sheets.items():
            nome = str(sheet_name)[:31]

            # Linha de início dos dados (após cabeçalho com logo)
            data_start_row = 1

            if logo_path:
                # Escreve com offset para deixar espaço para o logo
                data_start_row = 6
                df.to_excel(writer, index=False, sheet_name=nome,
                             startrow=data_start_row - 1)
            else:
                df.to_excel(writer, index=False, sheet_name=nome)

            ws = writer.sheets[nome]
            ws.sheet_view.showGridLines = False

            if logo_path:
                try:
                    # Insere logo
                    img = XLImage(logo_path)
                    img.width  = 80
                    img.height = 80
                    ws.add_image(img, "A1")
                except Exception:
                    pass

                # Título e subtítulo
                ws["C1"] = "KITAGAWA — Gestão de Abastecimento"
                ws["C1"].font = title_font
                ws["C2"] = f"Relatório: {nome}"
                ws["C2"].font = Font(color="D1FAE5", bold=True, size=12)
                ws["C3"] = f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}"
                ws["C3"].font = sub_font
                ws["C4"] = f"Total de registros: {len(df)}"
                ws["C4"].font = sub_font

                # Estilo cabeçalho da tabela
                header_row = data_start_row
                for ci, cell in enumerate(ws[header_row], 1):
                    cell.fill      = header_fill
                    cell.font      = header_font
                    cell.border    = thin
                    cell.alignment = Alignment(horizontal="center")
                    ws.column_dimensions[get_column_letter(ci)].width = 18

                # Estilo dados
                for ri in range(header_row + 1, header_row + len(df) + 1):
                    fill = dark_fill if ri % 2 == 0 else alt_fill
                    for cell in ws[ri]:
                        cell.fill      = fill
                        cell.font      = data_font
                        cell.border    = thin
                        cell.alignment = Alignment(horizontal="center")
            else:
                # Sem logo — estilo normal
                for ci, cell in enumerate(ws[1], 1):
                    cell.fill      = header_fill
                    cell.font      = header_font
                    cell.border    = thin
                    cell.alignment = Alignment(horizontal="center")
                    ws.column_dimensions[get_column_letter(ci)].width = 18
                for ri, row in enumerate(ws.iter_rows(min_row=2), 2):
                    fill = dark_fill if ri % 2 == 0 else alt_fill
                    for cell in row:
                        cell.fill      = fill
                        cell.font      = data_font
                        cell.border    = thin
                        cell.alignment = Alignment(horizontal="center")

    return buf.getvalue()


# ── PDF com logo ──────────────────────────────────────────────────
def df_to_pdf(df: pd.DataFrame, titulo: str,
              subtitulo: str = "", landscape: bool = True) -> bytes:
    from reportlab.lib.pagesizes import A4, landscape as rl_landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer, Image, HRFlowable)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_LEFT, TA_CENTER

    buf      = io.BytesIO()
    pagesize = rl_landscape(A4) if landscape else A4
    doc = SimpleDocTemplate(buf, pagesize=pagesize,
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)
    stls = getSampleStyleSheet()

    title_style = ParagraphStyle("T", parent=stls["Title"],
                                  fontSize=18, fontName="Helvetica-Bold",
                                  textColor=colors.HexColor("#10B981"),
                                  spaceAfter=2, alignment=TA_LEFT)
    sub_style   = ParagraphStyle("S", parent=stls["Normal"],
                                  fontSize=9,
                                  textColor=colors.HexColor("#64748B"),
                                  spaceAfter=10, alignment=TA_LEFT)

    elems = []

    # ── Cabeçalho com logo ────────────────────────────────────────
    logo_path = _get_logo_path()
    if logo_path:
        try:
            # Logo + título lado a lado numa tabela
            logo_img = Image(logo_path, width=2.2*cm, height=2.2*cm)
            header_data = [[
                logo_img,
                [Paragraph("KITAGAWA", title_style),
                 Paragraph("Sistema de Gestão de Abastecimento", sub_style),
                 Paragraph(titulo, ParagraphStyle("T2",
                     parent=stls["Normal"], fontSize=11,
                     textColor=colors.HexColor("#D1FAE5"), fontName="Helvetica-Bold")),
                 Paragraph(
                     subtitulo or
                     f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')} "
                     f"· {len(df)} registros",
                     sub_style)],
            ]]
            header_tbl = Table(header_data,
                               colWidths=[2.5*cm, pagesize[0]-5*cm])
            header_tbl.setStyle(TableStyle([
                ("VALIGN",      (0,0),(-1,-1), "MIDDLE"),
                ("LEFTPADDING", (0,0),(-1,-1), 0),
                ("RIGHTPADDING",(0,0),(-1,-1), 10),
            ]))
            elems.append(header_tbl)
        except Exception:
            # Fallback sem logo
            elems.append(Paragraph("KITAGAWA — Gestão de Abastecimento", title_style))
            elems.append(Paragraph(titulo, sub_style))
    else:
        elems.append(Paragraph("KITAGAWA — Gestão de Abastecimento", title_style))
        elems.append(Paragraph(titulo, sub_style))

    # Linha separadora
    elems.append(HRFlowable(width="100%", thickness=2,
                             color=colors.HexColor("#10B981"),
                             spaceAfter=10))

    # ── Tabela de dados ───────────────────────────────────────────
    cols  = df.columns.tolist()
    col_w = [(pagesize[0] - 3*cm) / len(cols)] * len(cols)

    # Limita a 500 linhas
    df_pdf = df.tail(500) if len(df) > 500 else df
    data   = [cols] + df_pdf.astype(str).values.tolist()

    tbl = Table(data, colWidths=col_w, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0,0),(-1,0), colors.HexColor("#10B981")),
        ("TEXTCOLOR",     (0,0),(-1,0), colors.white),
        ("FONTNAME",      (0,0),(-1,0), "Helvetica-Bold"),
        ("FONTSIZE",      (0,0),(-1,0), 8),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),
            [colors.HexColor("#1A1D23"), colors.HexColor("#111318")]),
        ("TEXTCOLOR",     (0,1),(-1,-1), colors.HexColor("#E2E8F0")),
        ("FONTSIZE",      (0,1),(-1,-1), 7),
        ("ALIGN",         (0,0),(-1,-1), "CENTER"),
        ("VALIGN",        (0,0),(-1,-1), "MIDDLE"),
        ("BOX",           (0,0),(-1,-1), 0.5, colors.HexColor("#2A2D35")),
        ("INNERGRID",     (0,0),(-1,-1), 0.2,  colors.HexColor("#2A2D35")),
        ("TOPPADDING",    (0,0),(-1,-1), 4),
        ("BOTTOMPADDING", (0,0),(-1,-1), 4),
    ]))
    elems.append(tbl)

    # Rodapé
    elems.append(Spacer(1, 0.3*cm))
    elems.append(HRFlowable(width="100%", thickness=0.5,
                             color=colors.HexColor("#1F3328")))
    elems.append(Paragraph(
        f"Kitagawa ERP v11 · {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        ParagraphStyle("foot", parent=stls["Normal"],
                       fontSize=7, textColor=colors.HexColor("#4A7C65"))))

    doc.build(elems)
    return buf.getvalue()


# ── Botões de download ────────────────────────────────────────────
def botoes_download(st, df: pd.DataFrame, nome_base: str,
                    titulo_pdf: str, colunas_pdf: list = None,
                    sheets_excel: dict = None):
    data_str = datetime.now().strftime("%Y%m%d")
    c1, c2, _ = st.columns([1, 1, 6])

    with c1:
        try:
            df_pdf = df[colunas_pdf].copy() if colunas_pdf else df.copy()
            pdf_bytes = df_to_pdf(df_pdf, titulo_pdf)
            st.download_button(
                "📄  PDF", pdf_bytes,
                f"{nome_base}_{data_str}.pdf",
                "application/pdf",
                use_container_width=True)
        except Exception as e:
            st.error(f"Erro PDF: {e}")

    with c2:
        try:
            xls_bytes = df_to_excel_multi(sheets_excel) if sheets_excel \
                        else df_to_excel(df)
            st.download_button(
                "📗  Excel", xls_bytes,
                f"{nome_base}_{data_str}.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True)
        except Exception as e:
            st.error(f"Erro Excel: {e}")
